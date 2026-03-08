from io import BytesIO
from datetime import datetime

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# =========================================================
# CONFIG
# =========================================================
st.set_page_config(
    page_title="Infrastructure BOQ Cost Report",
    page_icon="🏗️",
    layout="wide",
)

COMPANY_NAME = "INFINITY MAX CONTRACTING LLC, DUBAI"
AUTHOR_NAME = "Abdulla Zahin"
REPORT_VERSION = "BOQ Analyzer V8"


# =========================================================
# HELPERS
# =========================================================
def rupee(value):
    return f"₹{value:,.2f}"


def inject_css():
    st.markdown(
        """
        <style>
        .main-title {
            font-size: 2.2rem;
            font-weight: 700;
            margin-bottom: 0.2rem;
        }
        .sub-title {
            color: #666666;
            margin-bottom: 1rem;
        }
        .section-box {
            padding: 1rem;
            border-radius: 12px;
            background-color: #f5f7fa;
            margin-bottom: 1rem;
            border: 1px solid #e6e9ef;
        }
        .small-label {
            color: #666666;
            font-size: 0.9rem;
        }
        .footer-note {
            text-align: center;
            color: #666666;
            font-size: 0.9rem;
            padding-top: 1rem;
            padding-bottom: 1rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def validate_required_columns(df):
    required_columns = ["Item", "Category", "Quantity", "Unit Price"]
    return [col for col in required_columns if col not in df.columns]


def clean_numeric_data(df):
    df = df.copy()
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce")
    df["Unit Price"] = pd.to_numeric(df["Unit Price"], errors="coerce")
    return df


def run_boq_validation_checks(df):
    issues = []

    if df["Item"].astype(str).str.strip().eq("").any():
        issues.append("Some rows have blank Item names.")

    if df["Category"].astype(str).str.strip().eq("").any():
        issues.append("Some rows have blank Category values.")

    if df["Quantity"].isnull().any():
        issues.append("Some Quantity values are invalid or non-numeric.")

    if df["Unit Price"].isnull().any():
        issues.append("Some Unit Price values are invalid or non-numeric.")

    if (df["Quantity"] <= 0).any():
        issues.append("Some rows contain zero or negative Quantity values.")

    if (df["Unit Price"] < 0).any():
        issues.append("Some rows contain negative Unit Price values.")

    duplicate_rows = df.duplicated(subset=["Item", "Category"], keep=False)
    if duplicate_rows.any():
        issues.append("Possible duplicate Item + Category combinations found.")

    return issues


def analyze_boq(df):
    df = df.copy()
    df["Total Cost"] = df["Quantity"] * df["Unit Price"]

    total_project_cost = df["Total Cost"].sum()
    if total_project_cost <= 0:
        return None

    df["Cost Share (%)"] = (df["Total Cost"] / total_project_cost) * 100

    avg_cost = df["Total Cost"].mean()
    df["Cost Level"] = df["Total Cost"].apply(
        lambda x: "High" if x > avg_cost * 1.5 else "Normal"
    )

    category_summary = (
        df.groupby("Category", as_index=False)["Total Cost"]
        .sum()
        .sort_values(by="Total Cost", ascending=False)
    )

    category_summary["Category Share (%)"] = (
        category_summary["Total Cost"] / total_project_cost
    ) * 100

    most_expensive_item = df.loc[df["Total Cost"].idxmax()]
    top_5_items = df.sort_values(by="Total Cost", ascending=False).head(5).copy()
    high_cost_items = df[df["Cost Level"] == "High"].copy()

    top_category = category_summary.iloc[0]
    top_category_name = top_category["Category"]
    top_category_cost = top_category["Total Cost"]
    top_category_share = top_category["Category Share (%)"]

    if top_category_share >= 50:
        risk_level = "High"
        risk_message = (
            f"{top_category_share:.1f}% of total project cost comes from "
            f"{top_category_name}. This indicates strong cost concentration."
        )
    elif top_category_share >= 40:
        risk_level = "Moderate"
        risk_message = (
            f"{top_category_share:.1f}% of total project cost comes from "
            f"{top_category_name}. This suggests moderate cost concentration."
        )
    else:
        risk_level = "Low"
        risk_message = "Cost distribution appears reasonably balanced across categories."

    executive_summary = {
        "Total BOQ Items": len(df),
        "Total Estimated Cost": total_project_cost,
        "Number of Categories": df["Category"].nunique(),
        "Top Cost Category": top_category_name,
        "Top Cost Category Share (%)": top_category_share,
        "Most Expensive Item": most_expensive_item["Item"],
        "Most Expensive Item Cost": most_expensive_item["Total Cost"],
        "High-Cost Items Count": len(high_cost_items),
        "Cost Concentration Risk": risk_level,
        "Risk Note": risk_message,
    }

    return {
        "detailed_df": df,
        "category_summary": category_summary,
        "most_expensive_item": most_expensive_item,
        "top_5_items": top_5_items,
        "high_cost_items": high_cost_items,
        "total_project_cost": total_project_cost,
        "executive_summary": executive_summary,
        "top_category_cost": top_category_cost,
    }


def create_excel_report(results, project_info):
    output = BytesIO()

    detailed_df = results["detailed_df"].copy()
    category_summary = results["category_summary"].copy()
    top_5_items = results["top_5_items"].copy()
    high_cost_items = results["high_cost_items"].copy()
    total_project_cost = results["total_project_cost"]
    executive_summary = results["executive_summary"]

    boq_report_df = detailed_df[
        ["Item", "Category", "Quantity", "Unit Price", "Total Cost", "Cost Share (%)"]
    ].copy()

    current_date = datetime.now().strftime("%d-%m-%Y")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write main sheet starting at row 6 and column B
        sheet_name = "Detailed BOQ"
        boq_report_df.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=5,
            startcol=1,
            index=False
        )

        # Supporting sheets
        category_summary.to_excel(writer, sheet_name="Category Summary", index=False)
        top_5_items.to_excel(writer, sheet_name="Top Items", index=False)
        high_cost_items.to_excel(writer, sheet_name="High Cost Items", index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Styles
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        grey_fill = PatternFill("solid", fgColor="D9D9D9")
        light_fill = PatternFill("solid", fgColor="F2F2F2")

        # Company title
        worksheet.merge_cells("B1:G1")
        worksheet["B1"] = COMPANY_NAME
        worksheet["B1"].font = title_font
        worksheet["B1"].alignment = center_align
        worksheet["B1"].fill = light_fill
        worksheet["B1"].border = border

        # Header info
        worksheet["B2"] = "Project Name"
        worksheet["C2"] = project_info["project_name"]

        worksheet["B3"] = "Client Name"
        worksheet["C3"] = project_info["client_name"]

        worksheet["B4"] = "Date"
        worksheet["C4"] = current_date

        worksheet["E2"] = "Prepared By"
        worksheet["F2"] = AUTHOR_NAME

        worksheet["E3"] = "Project Location"
        worksheet["F3"] = project_info["project_location"]

        worksheet["E4"] = "Total Cost"
        worksheet["F4"] = total_project_cost

        for cell in ["B2", "B3", "B4", "E2", "E3", "E4"]:
            worksheet[cell].font = bold_font
            worksheet[cell].fill = light_fill
            worksheet[cell].border = border
            worksheet[cell].alignment = left_align

        for cell in ["C2", "C3", "C4", "F2", "F3", "F4"]:
            worksheet[cell].border = border
            worksheet[cell].alignment = left_align

        worksheet["F4"].number_format = '#,##0.00'

        # BOQ title row
        worksheet.merge_cells("B5:G5")
        worksheet["B5"] = "Bill of Quantities"
        worksheet["B5"].font = bold_font
        worksheet["B5"].alignment = center_align
        worksheet["B5"].fill = grey_fill
        worksheet["B5"].border = border

        # Header row formatting
        header_row = 6
        for col in range(2, 8):  # B:G
            cell = worksheet.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = grey_fill
            cell.alignment = center_align
            cell.border = border

        # Data area formatting
        final_data_row = 6 + len(boq_report_df)
        for row in range(7, final_data_row + 1):
            for col in range(2, 8):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_align

        # Numeric formatting
        for row in range(7, final_data_row + 1):
            worksheet[f"D{row}"].number_format = '0.00'
            worksheet[f"E{row}"].number_format = '#,##0.00'
            worksheet[f"F{row}"].number_format = '#,##0.00'
            worksheet[f"G{row}"].number_format = '0.00'

        # Signature section
        signature_label_row = final_data_row + 3
        signature_name_row = final_data_row + 4

        worksheet.merge_cells(f"B{signature_label_row}:G{signature_label_row}")
        worksheet[f"B{signature_label_row}"] = "Prepared By / Signature"
        worksheet[f"B{signature_label_row}"].font = bold_font
        worksheet[f"B{signature_label_row}"].alignment = center_align
        worksheet[f"B{signature_label_row}"].border = border

        worksheet.merge_cells(f"B{signature_name_row}:G{signature_name_row + 1}")
        worksheet[f"B{signature_name_row}"] = AUTHOR_NAME
        worksheet[f"B{signature_name_row}"].alignment = center_align
        worksheet[f"B{signature_name_row}"].border = border

        # Column widths
        column_widths = {
            "A": 3,
            "B": 22,
            "C": 16,
            "D": 12,
            "E": 12,
            "F": 14,
            "G": 14,
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    output.seek(0)
    return output


# =========================================================
# UI
# =========================================================
inject_css()

st.markdown(
    '<div class="main-title">Infrastructure BOQ Cost Report</div>',
    unsafe_allow_html=True,
)
st.markdown(
    '<div class="sub-title">Engineering Cost Review Dashboard</div>',
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("📋 Project Information")

    project_name = st.text_input("Project Name", value="Infrastructure BOQ Review")
    project_location = st.text_input("Project Location", value="Dubai")
    client_name = st.text_input("Client Name", value="Internal / Not Specified")

    st.markdown("---")
    st.header("📄 Required CSV Columns")
    st.write("Item")
    st.write("Category")
    st.write("Quantity")
    st.write("Unit Price")

uploaded_file = st.file_uploader("Upload BOQ CSV", type=["csv"])

if uploaded_file is None:
    st.info("Upload a BOQ CSV file to generate the report.")
    st.stop()

try:
    raw_df = pd.read_csv(uploaded_file)
except Exception as e:
    st.error(f"Error reading file: {e}")
    st.stop()

missing_columns = validate_required_columns(raw_df)
if missing_columns:
    st.error(f"Missing required columns: {', '.join(missing_columns)}")
    st.stop()

df = clean_numeric_data(raw_df)

validation_issues = run_boq_validation_checks(df)

results = analyze_boq(df)
if results is None:
    st.error("Total project cost is zero or invalid. Please check your BOQ data.")
    st.stop()

project_info = {
    "project_name": project_name,
    "project_location": project_location,
    "client_name": client_name,
}

detailed_df = results["detailed_df"]
category_summary = results["category_summary"]
most_expensive_item = results["most_expensive_item"]
top_5_items = results["top_5_items"]
high_cost_items = results["high_cost_items"]
executive_summary = results["executive_summary"]
total_project_cost = results["total_project_cost"]

# =========================================================
# PROJECT OVERVIEW
# =========================================================
st.subheader("Project Overview")

meta1, meta2, meta3 = st.columns(3)
meta1.metric("Total Project Cost", rupee(total_project_cost))
meta2.metric("BOQ Items", executive_summary["Total BOQ Items"])
meta3.metric("Categories", executive_summary["Number of Categories"])

# =========================================================
# EXECUTIVE SUMMARY
# =========================================================
st.subheader("Executive Summary")

st.markdown(
    f"""
    <div class="section-box">
    <b>Company:</b> {COMPANY_NAME}<br><br>
    <b>Project Name:</b> {project_name}<br><br>
    <b>Project Location:</b> {project_location}<br><br>
    <b>Client Name:</b> {client_name}<br><br>
    <b>Prepared By:</b> {AUTHOR_NAME}<br><br>
    <b>Top Cost Category:</b> {executive_summary["Top Cost Category"]} 
    ({executive_summary["Top Cost Category Share (%)"]:.2f}%)<br><br>
    <b>Most Expensive Item:</b> {executive_summary["Most Expensive Item"]} 
    ({rupee(executive_summary["Most Expensive Item Cost"])})<br><br>
    <b>Cost Concentration Risk:</b> {executive_summary["Cost Concentration Risk"]}<br>
    <b>Risk Note:</b> {executive_summary["Risk Note"]}
    </div>
    """,
    unsafe_allow_html=True,
)

# =========================================================
# VALIDATION CHECKS
# =========================================================
st.subheader("BOQ Validation Checks")

if validation_issues:
    for issue in validation_issues:
        st.warning(issue)
else:
    st.success("No major BOQ validation issues detected.")

# =========================================================
# CATEGORY ANALYSIS
# =========================================================
st.subheader("Cost Distribution")

chart_col1, chart_col2 = st.columns(2)

with chart_col1:
    fig_bar, ax_bar = plt.subplots(figsize=(8, 5))
    ax_bar.bar(category_summary["Category"], category_summary["Total Cost"])
    ax_bar.set_title("Cost by Category")
    ax_bar.set_xlabel("Category")
    ax_bar.set_ylabel("Total Cost")
    plt.xticks(rotation=45)
    plt.tight_layout()
    st.pyplot(fig_bar)

with chart_col2:
    fig_pie, ax_pie = plt.subplots(figsize=(7, 7))
    ax_pie.pie(
        category_summary["Total Cost"],
        labels=category_summary["Category"],
        autopct="%1.1f%%",
    )
    ax_pie.set_title("Category Share")
    st.pyplot(fig_pie)

# =========================================================
# KEY COST DRIVERS
# =========================================================
st.subheader("Key Cost Drivers")

st.write("Most Expensive Item")
most_expensive_display = pd.DataFrame([most_expensive_item])[
    ["Item", "Category", "Quantity", "Unit Price", "Total Cost", "Cost Share (%)", "Cost Level"]
]
st.dataframe(most_expensive_display, use_container_width=True)

st.write("Top 5 Cost Drivers")
st.dataframe(
    top_5_items[
        ["Item", "Category", "Quantity", "Unit Price", "Total Cost", "Cost Share (%)", "Cost Level"]
    ],
    use_container_width=True,
)

# =========================================================
# COST ALERTS
# =========================================================
st.subheader("Cost Alerts")

if executive_summary["Cost Concentration Risk"] == "High":
    st.error(executive_summary["Risk Note"])
elif executive_summary["Cost Concentration Risk"] == "Moderate":
    st.warning(executive_summary["Risk Note"])
else:
    st.success(executive_summary["Risk Note"])

if high_cost_items.empty:
    st.info("No unusually high-cost items detected.")
else:
    st.dataframe(
        high_cost_items[
            ["Item", "Category", "Quantity", "Unit Price", "Total Cost", "Cost Share (%)", "Cost Level"]
        ],
        use_container_width=True,
    )

# =========================================================
# BOQ REVIEW
# =========================================================
st.subheader("BOQ Item Review")

search_term = st.text_input("Search Item")
category_filter = st.selectbox(
    "Filter Category",
    ["All"] + sorted(detailed_df["Category"].dropna().unique().tolist()),
)

filtered_df = detailed_df.copy()

if search_term:
    filtered_df = filtered_df[
        filtered_df["Item"].astype(str).str.contains(search_term, case=False, na=False)
    ]

if category_filter != "All":
    filtered_df = filtered_df[filtered_df["Category"] == category_filter]

st.dataframe(
    filtered_df[
        ["Item", "Category", "Quantity", "Unit Price", "Total Cost", "Cost Share (%)", "Cost Level"]
    ],
    use_container_width=True,
)

# =========================================================
# EXPORT
# =========================================================
st.subheader("Export Report")

excel_data = create_excel_report(results, project_info)
st.download_button(
    label="Download Excel Report",
    data=excel_data,
    file_name="boq_cost_report_v8.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# =========================================================
# FOOTER
# =========================================================
st.markdown("---")
st.markdown(
    f'<div class="footer-note">Prepared by {AUTHOR_NAME} | {COMPANY_NAME} | {REPORT_VERSION}</div>',
    unsafe_allow_html=True,
)